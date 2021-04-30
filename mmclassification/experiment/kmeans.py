import numpy as np
import os
import pickle
import torch
import openpyxl

def cosine_dist(feature, center, norm):
    if norm:
        return - torch.matmul(feature/feature.norm(dim=1, keepdim=True), (center/center.norm(dim=1, keepdim=True)).T)
    else:
        return - torch.matmul(feature, center.T)


def mse_dist(feature, center, norm):
    if norm:
        dist = (feature/feature.norm(dim=1, keepdim=True)).unsqueeze(1) - (center/center.norm(dim=1, keepdim=True)).unsqueeze(0)
    else:
        dist = feature.unsqueeze(1) - center.unsqueeze(0)
    return dist.pow(2).sum(2)

def k_means(mode, norm, features, center, gt_label):
    modes = {'cosine': cosine_dist, 'mse': mse_dist}
    dist_fn = modes.get(mode)
    center_now = center
    feat = features
    refs = torch.Tensor(range(31)).unsqueeze(1)
    gt_mask = (gt_label == refs).unsqueeze(2)
    gt_class_count = gt_mask.sum(dim=1)

    dist = dist_fn(feat, center_now, norm)
    _, pred_init = torch.min(dist, dim=1)

    cluster_iter = 0

    while True:

        dist = dist_fn(feat, center_now, norm)
        _, pred = torch.min(dist, dim=1)

        if cluster_iter == 0:
            result_init = pred == gt_label
            classwise_result_init = result_init.unsqueeze(0) * gt_mask.squeeze()
            acc_init = result_init.sum()/len(gt_label)
            classwise_acc_init = (classwise_result_init.sum(dim=1, keepdim=True)/gt_class_count).squeeze()
        
        if (cluster_iter != 0 and sum(pred == pred_init) == len(pred)) or cluster_iter == 100:
            result = pred == gt_label
            classwise_result = result.unsqueeze(0) * gt_mask.squeeze()
            acc = result.sum()/len(gt_label)
            classwise_acc = (classwise_result.sum(dim=1, keepdim=True)/gt_class_count).squeeze()
            return acc_init, classwise_acc_init, acc, classwise_acc
    
        mask = (pred == refs).unsqueeze(2)
        num_mask = mask.sum(dim=1)
        center_now = torch.sum(feat.unsqueeze(0) * mask, dim=1) / (num_mask + 10e-6)
        pred_init = pred
        cluster_iter += 1

def calculate_center(feature, label):
    refs = torch.Tensor(range(31)).unsqueeze(1)
    mask = (label == refs).unsqueeze(2)
    num = mask.sum(dim=1)
    feat = feature.unsqueeze(0)
    center = torch.sum(feat * mask, dim=1) / (num + 10e-6)

    return center

def main():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.tilte='classwise_office'

    label_names = sorted(os.listdir('/lustre/S/wangyu/dataset/office-31/amazon/images/'))
    worksheet.cell(2, 1, 'domain')
    worksheet.cell(2, 2, 'distance_algo')
    worksheet.cell(2, 3, 'top-1')
    for idx, label_name in enumerate(label_names):
        worksheet.cell(2, idx + 4, label_name)

    line = 3

    features_t = torch.from_numpy(np.vstack(pickle.load(open('features_t.pkl', 'rb'))))
    mlp_features_t = torch.from_numpy(np.vstack(pickle.load(open('mlp_features_t.pkl', 'rb'))))
    results_t = torch.from_numpy(np.vstack(pickle.load(open('results_t.pkl', 'rb'))))
    label_t = torch.from_numpy(pickle.load(open('gt_labels_t.pkl', 'rb')))
    features_s = torch.from_numpy(np.vstack(pickle.load(open('features_s.pkl', 'rb'))))
    mlp_features_s = torch.from_numpy(np.vstack(pickle.load(open('mlp_features_s.pkl', 'rb'))))
    results_s = torch.from_numpy(np.vstack(pickle.load(open('results_s.pkl', 'rb'))))
    label_s = torch.from_numpy(pickle.load(open('gt_labels_s.pkl', 'rb')))
    
    
    modes = ['cosine', 'mse']
    norm_mode = [True, False]

    cluster_modes = []
    for m in modes:
        for n in norm_mode:
            mode = {'mode': m, 'norm': n}
            cluster_modes.append(mode)

    center_s = calculate_center(features_s, label_s)
    mlp_center_s = calculate_center(mlp_features_s, label_s)
    
    for mode in cluster_modes:
        acc_init, classwise_init, acc, classwise = k_means(features=features_s, center=center_s, gt_label=label_s, **mode)
        norm_ab = 'norm' if mode['norm'] else 'unnorm'
        domain = 'source'
        mode_ab = mode['mode'] + '_' + norm_ab
        worksheet.cell(line, 1, domain)
        worksheet.cell(line, 2, mode_ab)
        worksheet.cell(line, 3, acc.item())
        for idx, i in enumerate(classwise):
            worksheet.cell(line, idx + 4, i.item())
        worksheet.cell(line + 1, 2, 'initial source center')
        worksheet.cell(line + 1, 3, acc_init.item())
        for idx, i in enumerate(classwise_init):
            worksheet.cell(line + 1, idx + 4, i.item())

        line += 2
        
    center_t = calculate_center(features_t, label_t)
    mlp_center_t = calculate_center(mlp_features_t, label_t)

    for mode in cluster_modes:
        acc_init, classwise_init, acc, classwise = k_means(features=features_t, center=center_s, gt_label=label_t, **mode)
        norm_ab = 'norm' if mode['norm'] else 'unnorm'
        domain = 'target'
        mode_ab = mode['mode'] + '_' + norm_ab
        worksheet.cell(line, 1, domain)
        worksheet.cell(line, 2, mode_ab)
        worksheet.cell(line, 3, acc.item())
        for idx, i in enumerate(classwise):
            worksheet.cell(line, idx + 4, i.item())
        worksheet.cell(line + 1, 2, 'initial source center')
        worksheet.cell(line + 1, 3, acc_init.item())
        for idx, i in enumerate(classwise_init):
            worksheet.cell(line + 1, idx + 4, i.item())
        line += 2

    workbook.save(filename='cluster_result_resnet50.xlsx')

if __name__ == '__main__':
    main()
