import numpy as np
import os
import pickle
import torch
import openpyxl

device = torch.device('cuda')
refs = torch.Tensor(range(31)).unsqueeze(1).to(device)

def cosine_dist(feature, center, norm):
    if norm:
        return - torch.matmul(feature/feature.norm(dim=1, keepdim=True), (center/center.norm(dim=1, keepdim=True)).T)
    else:
        return - torch.matmul(feature, center.T)


def mse_dist(feature, center, norm):
    if norm:
        dist = (feature/feature.norm(dim=1, keepdim=True)).unsqueeze(1) - (center/center.norm(dim=1, keepdim=True)).unsqueeze(0)
    else:
        dist = feature.unsqueeze(1) - center.unsqueeze(0)
    return dist.pow(2).sum(2)

def k_means(mode, norm, features, center, gt_label):
    global refs

    modes = {'cosine': cosine_dist, 'mse': mse_dist}
    dist_fn = modes.get(mode)
    center_now = center
    feat = features
    gt_mask = (gt_label == refs).unsqueeze(2)
    gt_class_count = gt_mask.sum(dim=1)

    dist = dist_fn(feat, center_now, norm)
    _, pred_init = torch.min(dist, dim=1)

    cluster_iter = 0

    while True:

        dist = dist_fn(feat, center_now, norm)
        _, pred = torch.min(dist, dim=1)

        if cluster_iter == 0:
            result_init = pred == gt_label
            classwise_result_init = result_init.unsqueeze(0) * gt_mask.squeeze()
            acc_init = result_init.sum()/len(gt_label)
            classwise_acc_init = (classwise_result_init.sum(dim=1, keepdim=True)/gt_class_count).squeeze()

        consist = (pred != pred_init).float()
        convergence = torch.matmul(consist, consist.T)
        
        if ((cluster_iter != 0) and convergence == 0) or cluster_iter == 100:
            result = pred == gt_label
            classwise_result = result.unsqueeze(0) * gt_mask.squeeze()
            acc = result.sum()/len(gt_label)
            classwise_acc = (classwise_result.sum(dim=1, keepdim=True)/gt_class_count).squeeze()
            return acc_init.cpu(), classwise_acc_init.cpu(), acc.cpu(), classwise_acc.cpu()
    
        mask = (pred == refs).unsqueeze(2)
        num_mask = mask.sum(dim=1)
        center_now = torch.sum(feat.unsqueeze(0) * mask, dim=1) / (num_mask + 10e-6)
        pred_init = pred
        cluster_iter += 1

def calculate_center(feature, label):
    global refs
    mask = (label == refs).unsqueeze(2)
    num = mask.sum(dim=1)
    feat = feature.unsqueeze(0)
    center = torch.sum(feat * mask, dim=1) / (num + 10e-6)

    return center

def main():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.tilte='classwise_office'

    label_names = os.listdir('/lustre/S/wangyu/dataset/office-31/amazon/images/')
    CLASSES = np.array([
        'back_pack', 'bike', 'bike_helmet', 'bookcase', 'bottle', 'calculator', 'desk_chair',
        'desk_lamp', 'desktop_computer', 'file_cabinet', 'headphones', 'keyboard', 'laptop_computer',
        'letter_tray', 'mobile_phone', 'monitor', 'mouse', 'mug', 'paper_notebook', 'pen', 'phone',
        'printer', 'projector', 'punchers', 'ring_binder', 'ruler', 'scissors', 'speaker', 'stapler',
        'tape_dispenser', 'trash_can'
    ])
    """
    CLASSES = [
        'aeroplane', 'bicycle', 'bus', 'car', 'horse', 'knife', 'motorcycle', 'person',
        'plant', 'skateboard', 'train', 'truck'
    ]
    """
    worksheet.cell(2, 1, 'domain')
    worksheet.cell(2, 2, 'distance_algo')
    worksheet.cell(2, 3, 'top-1')
    for idx, label_name in enumerate(CLASSES):
        worksheet.cell(2, idx + 4, label_name)

    line = 3

    features_t = torch.from_numpy(np.vstack(pickle.load(open('features_t.pkl', 'rb')))).to(device)
    mlp_features_t = torch.from_numpy(np.vstack(pickle.load(open('mlp_features_t.pkl', 'rb')))).to(device)
    features_s = torch.from_numpy(np.vstack(pickle.load(open('features_s.pkl', 'rb')))).to(device)
    mlp_features_s = torch.from_numpy(np.vstack(pickle.load(open('mlp_features_s.pkl', 'rb')))).to(device)
    label_s = torch.from_numpy(pickle.load(open('gt_labels_s.pkl', 'rb'))).to(device)
    label_t = torch.from_numpy(pickle.load(open('gt_labels_t.pkl', 'rb'))).to(device)

    """
    CLASSES = np.array([
        'back_pack', 'bike', 'bike_helmet', 'bookcase', 'bottle', 'calculator', 'desk_chair',
        'desk_lamp', 'desktop_computer', 'file_cabinet', 'headphones', 'keyboard', 'laptop_computer',
        'letter_tray', 'mobile_phone', 'monitor', 'mouse', 'mug', 'paper_notebook', 'pen', 'phone',
        'printer', 'projector', 'punchers', 'ring_binder', 'ruler', 'scissors', 'speaker', 'stapler',
        'tape_dispenser', 'trash_can'
    ])

    label_s_trans = torch.zeros_like(label_s)
    for idx, i in enumerate(label_s):
        label_s_trans[idx] = np.where(CLASSES == label_names[i])[0].item()
    """

    modes = ['cosine', 'mse']
    norm_mode = [True, False]

    cluster_modes = []
    for m in modes:
        for n in norm_mode:
            mode = {'mode': m, 'norm': n}
            cluster_modes.append(mode)

    center_s = calculate_center(features_s, label_s)
    mlp_center_s = calculate_center(mlp_features_s, label_s)
    
    for mode in cluster_modes:
        acc_init, classwise_init, acc, classwise = k_means(features=features_s, center=center_s, gt_label=label_s, **mode)
        mlp_acc_init, mlp_classwise_init, mlp_acc, mlp_classwise = k_means(features=mlp_features_s, center=mlp_center_s, gt_label=label_s, **mode)
        norm_ab = 'norm' if mode['norm'] else 'unnorm'
        domain = 'source'
        mode_ab = mode['mode'] + '_' + norm_ab
        worksheet.cell(line, 1, domain)
        worksheet.cell(line, 2, mode_ab)
        worksheet.cell(line, 3, acc.item())
        for idx, i in enumerate(classwise):
            worksheet.cell(line, idx + 4, i.item())
        worksheet.cell(line + 1, 2, 'initial source center')
        worksheet.cell(line + 1, 3, acc_init.item())
        for idx, i in enumerate(classwise_init):
            worksheet.cell(line + 1, idx + 4, i.item())
        worksheet.cell(line + 2, 2, 'mlp')
        worksheet.cell(line + 2, 3, mlp_acc.item())
        for idx, i in enumerate(mlp_classwise):
            worksheet.cell(line + 2, idx + 4, i.item())
        worksheet.cell(line + 3, 2, 'initial source center')
        worksheet.cell(line + 3, 3, mlp_acc_init.item())
        for idx, i in enumerate(mlp_classwise_init):
            worksheet.cell(line + 3, idx + 4, i.item())

        line += 4
        
    center_t = calculate_center(features_t, label_t)
    mlp_center_t = calculate_center(mlp_features_t, label_t)
    line += 1 
    for mode in cluster_modes:
        acc_init, classwise_init, acc, classwise = k_means(features=features_t, center=center_s, gt_label=label_t, **mode)
        mlp_acc_init, mlp_classwise_init, mlp_acc, mlp_classwise = k_means(features=mlp_features_t, center=mlp_center_s, gt_label=label_t, **mode)
        norm_ab = 'norm' if mode['norm'] else 'unnorm'
        domain = 'target'
        mode_ab = mode['mode'] + '_' + norm_ab
        worksheet.cell(line, 1, domain)
        worksheet.cell(line, 2, mode_ab)
        worksheet.cell(line, 3, acc.item())
        for idx, i in enumerate(classwise):
            worksheet.cell(line, idx + 4, i.item())
        worksheet.cell(line + 1, 2, 'initial source center')
        worksheet.cell(line + 1, 3, acc_init.item())
        for idx, i in enumerate(classwise_init):
            worksheet.cell(line + 1, idx + 4, i.item())
        worksheet.cell(line + 2, 2, 'mlp')
        worksheet.cell(line + 2, 3, mlp_acc.item())
        for idx, i in enumerate(mlp_classwise):
            worksheet.cell(line + 2, idx + 4, i.item())
        worksheet.cell(line + 3, 2, 'initial source center')
        worksheet.cell(line + 3, 3, mlp_acc_init.item())
        for idx, i in enumerate(mlp_classwise_init):
            worksheet.cell(line + 3, idx + 4, i.item())

        line += 4

    line += 1
    for mode in cluster_modes:
        acc_init, classwise_init, acc, classwise = k_means(features=features_t, center=center_t, gt_label=label_t, **mode)
        mlp_acc_init, mlp_classwise_init, mlp_acc, mlp_classwise = k_means(features=mlp_features_t, center=mlp_center_t, gt_label=label_t, **mode)
        norm_ab = 'norm' if mode['norm'] else 'unnorm'
        domain = 'target'
        mode_ab = mode['mode'] + '_' + norm_ab
        worksheet.cell(line, 1, domain)
        worksheet.cell(line, 2, mode_ab)
        worksheet.cell(line, 3, acc.item())
        for idx, i in enumerate(classwise):
            worksheet.cell(line, idx + 4, i.item())
        worksheet.cell(line + 1, 2, 'initial target center')
        worksheet.cell(line + 1, 3, acc_init.item())
        for idx, i in enumerate(classwise_init):
            worksheet.cell(line + 1, idx + 4, i.item())
        worksheet.cell(line + 2, 2, 'mlp')
        worksheet.cell(line + 2, 3, mlp_acc.item())
        for idx, i in enumerate(mlp_classwise):
            worksheet.cell(line + 2, idx + 4, i.item())
        worksheet.cell(line + 3, 2, 'initial target center')
        worksheet.cell(line + 3, 3, mlp_acc_init.item())
        for idx, i in enumerate(mlp_classwise_init):
            worksheet.cell(line + 3, idx + 4, i.item())

        line += 4



    workbook.save(filename='cluster_result_source-fc-2048-wa.xlsx')

if __name__ == '__main__':
    main()
